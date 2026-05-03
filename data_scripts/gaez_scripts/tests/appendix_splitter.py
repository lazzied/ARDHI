"""
Scratch script for splitting appendix-style source files during GAEZ data prep.

Script to split soil requirements Excel file into separate appendices
Creates 3 files: Appendix_6-3.xlsx, Appendix_6-4.xlsx, Appendix_6-5.xlsx
Each file includes a README sheet with specific documentation
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

# README content for each appendix
README_CONTENT = {
    '6-3': {
        'title': 'Soil Requirements for Rain-fed Crops',
        'intro': 'Soil requirements are presented for 45 crops producing under rain-fed conditions, assuming high, intermediate and low levels of inputs and management conditions\nSoil requirements ratings are presented separately for Chemical and Physical Soil Characteristics, Soil Textures, Soil Drainage and Soil Phases.',
        'sections': [
            {
                'header': 'Soil profile',
                'content': 'Soil characteristics suitability ratings are empirical coefficients that reflect the effect the value of the soil characteristic has on the yield potential of a specific crop.\nS0 - No constraint (100)       S2 - Moderate (70)                   S4  - Very severe constraint (30)\nS1 - Slight constraint (90)    S3 - Severe constraint (50)   N -    Not suitable (10)'
            },
            {
                'header': 'Texture',
                'content': 'Soil texture conditions are influencing various soil qualities (SQ1, SQ2, SQ3 and SQ7). Soil workability ratings differ for high, intermediate and low inputs. Soil texture ratings are compiled for 13 texture classes.'
            },
            {
                'header': 'Drainage',
                'content': 'Soil drainage ratings are varying by crop and may vary by prevalent soil texture conditions. Assumptions for artificial soil drainage differ by input levels. High level inputs assume full and adequate artificial drainage systems are installed while low and intermediate inputs assume no artificial drainage.'
            },
            {
                'header': 'Soil phase',
                'content': 'The soil phase ratings have been compiled by input level (high, intermediate and low). The ratings represent constraints implied by the occurrence of soil phases in percentage (100% rating no constraint to 0% rendering a soil totally unsuitable). The soil phases are organized by soil quality to which they apply and by level of input and management and water supply system. Two rating types have been used: (i) Soil phase rating is applying to 100% of the extent of the soil unit to which the soil phase is attributed and (ii) soil phase rating is assumed to affect 50% of the soil to which it is attributed while the other 50% is assumed not to be affected.'
            }
        ]
    },
    '6-4': {
        'title': 'Soil Requirements for Irrigated Crops (Gravity Irrigation)',
        'intro': 'Soil requirements are presented for 45 crops producing under gravity irrigation systems, assuming high and intermediate levels of inputs and management conditions\nSoil requirements ratings are presented separately for Chemical and Physical Soil Characteristics, Soil Textures, Soil Drainage and Soil Phases.',
        'sections': [
            {
                'header': 'Soil profile',
                'content': 'Soil characteristics suitability ratings are empirical coefficients that reflect the effect the value of the soil characteristic has on the yield potential of a specific crop.\nS0 - No constraint (100)        S2 - Moderate (70)                   S4  - Very severe constraint (30)\nS1 - Slight constraint (90)      S3 - Severe constraint (50)         N - Not suitable (10)'
            },
            {
                'header': 'Texture',
                'content': 'Soil texture conditions are influencing various soil qualities (SQ1, SQ2, SQ3 and SQ7). Soil workability ratings differ for high and intermediate inputs. Soil texture ratings are compiled for 13 texture classes.'
            },
            {
                'header': 'Drainage',
                'content': 'Soil drainage ratings are varying by crop and may vary by prevalent soil texture conditions. Assumptions for artificial soil drainage differ by input levels. High level inputs assume full and adequate artificial drainage systems are installed while intermediate inputs assume no artificial drainage.'
            },
            {
                'header': 'Soil phase',
                'content': 'The soil phase ratings have been compiled by input level (high and intermediate). The ratings represent constraints implied by the occurrence of soil phases in percentage (100% rating no constraint to 0% rendering a soil totally unsuitable). The soil phases are organized by soil quality to which they apply and by level of input and management and water supply system. Two rating types have been used: (i) Soil phase rating is applying to 100% of the extent of the soil unit to which the soil phase is attributed and (ii) soil phase rating is assumed to affect 50% of the soil to which it is attributed while the other 50% is assumed not to be affected.'
            }
        ]
    },
    '6-5': {
        'title': 'Soil Requirements for Irrigated Crops (Drip Irrigation)',
        'intro': 'Soil requirements are presented for 15 crops producing under drip irrigation systems, assuming high and intermediate levels of inputs and management conditions\nSoil requirements ratings are presented separately for Chemical and Physical Soil Characteristics, Soil Textures, Soil Drainage and Soil Phases.',
        'sections': [
            {
                'header': 'Soil profile',
                'content': 'Soil characteristics suitability ratings are empirical coefficients that reflect the effect the value of the soil characteristic has on the yield potential of a specific crop.\nS0 - No constraint (100)        S2 - Moderate (70)                     S4  - Very severe constraint (30)\nS1 - Slight constraint (90)    S3 - Severe constraint (50)        N - Not suitable (10)'
            },
            {
                'header': 'Texture',
                'content': 'Soil texture conditions are influencing various soil qualities (SQ1, SQ2, SQ3 and SQ7). Soil workability ratings differ for high and intermediate inputs. Soil texture ratings are compiled for 13 texture classes.'
            },
            {
                'header': 'Drainage',
                'content': 'Soil drainage ratings are varying by crop and may vary by prevalent soil texture conditions. Assumptions for artificial soil drainage differ by input levels. High level inputs assume full and adequate artificial drainage systems are installed while intermediate inputs assume no artificial drainage.'
            },
            {
                'header': 'Soil phase',
                'content': 'The soil phase ratings have been compiled by input level (high and intermediate). The ratings represent constraints implied by the occurrence of soil phases in percentage (100% rating no constraint to 0% rendering a soil totally unsuitable). The soil phases are organized by soil quality to which they apply and by level of input and management and water supply system. Two rating types have been used: (i) Soil phase rating is applying to 100% of the extent of the soil unit to which the soil phase is attributed and (ii) soil phase rating is assumed to affect 50% of the soil to which it is attributed while the other 50% is assumed not to be affected.'
            }
        ]
    }
}

# Sheet patterns for each appendix
SHEET_PATTERNS = {
    '6-3': ['A6-3.1', 'A6-3.2', 'A6-3.3', 'A6-3.4', 'A6-3.5', 'A6-3.6'],
    '6-4': ['A6-4.1', 'A6-4.2', 'A6-4.3', 'A6-4.4', 'A6-4.5'],
    '6-5': ['A6-5.1', 'A6-5.2', 'A6-5.3', 'A6-5.4', 'A6-5.5']
}


def create_readme_sheet(wb, appendix_key):
    """Create a formatted README sheet for the workbook"""
    ws = wb.create_sheet("README", 0)  # Insert as first sheet
    
    readme = README_CONTENT[appendix_key]
    
    # Title styling
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    normal_font = Font(name='Arial', size=11)
    
    # Set column width
    ws.column_dimensions['A'].width = 100
    
    row = 1
    
    # Title
    ws.cell(row=row, column=1, value=readme['title'])
    ws.cell(row=row, column=1).font = title_font
    ws.cell(row=row, column=1).fill = title_fill
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 30
    row += 2
    
    # Introduction
    ws.cell(row=row, column=1, value=readme['intro'])
    ws.cell(row=row, column=1).font = normal_font
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 45
    row += 2
    
    # Sections
    for section in readme['sections']:
        # Section header
        ws.cell(row=row, column=1, value=section['header'])
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Section content
        ws.cell(row=row, column=1, value=section['content'])
        ws.cell(row=row, column=1).font = normal_font
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Calculate row height based on content
        lines = section['content'].count('\n') + 1
        ws.row_dimensions[row].height = max(30, lines * 15)
        row += 2
    
    return ws


def split_excel_file(input_file, output_dir='/mnt/user-data/outputs'):
    """
    Split the main Excel file into 3 separate files based on appendix patterns
    
    Args:
        input_file: Path to the original Excel file
        output_dir: Directory to save the split files
    """
    
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"Loading workbook: {input_file}")
    wb_source = openpyxl.load_workbook(input_file, data_only=False)
    
    # Get all sheet names
    all_sheets = wb_source.sheetnames
    print(f"Found {len(all_sheets)} sheets in source file")
    
    # Process each appendix
    for appendix_key, patterns in SHEET_PATTERNS.items():
        print(f"\n{'='*60}")
        print(f"Processing Appendix {appendix_key}")
        print(f"{'='*60}")
        
        # Create new workbook
        wb_new = openpyxl.Workbook()
        wb_new.remove(wb_new.active)  # Remove default sheet
        
        # Add README sheet
        print("Creating README sheet...")
        create_readme_sheet(wb_new, appendix_key)
        
        # Find and copy matching sheets
        sheets_copied = 0
        for pattern in patterns:
            # Find sheet that contains this pattern
            matching_sheets = [s for s in all_sheets if pattern in s]
            
            if matching_sheets:
                source_sheet_name = matching_sheets[0]
                print(f"  Copying sheet: {source_sheet_name}")
                
                # Copy sheet from source to new workbook
                source_sheet = wb_source[source_sheet_name]
                target_sheet = wb_new.create_sheet(source_sheet_name)
                
                # Copy all cells, formulas, and formatting
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet[cell.coordinate]
                        
                        # Copy value or formula
                        if cell.value is not None:
                            if cell.data_type == 'f':
                                target_cell.value = f'={cell.value}'
                            else:
                                target_cell.value = cell.value
                        
                        # Copy formatting
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.protection = cell.protection.copy()
                            target_cell.alignment = cell.alignment.copy()
                
                # Copy column widths
                for col_letter in source_sheet.column_dimensions:
                    if col_letter in source_sheet.column_dimensions:
                        target_sheet.column_dimensions[col_letter].width = \
                            source_sheet.column_dimensions[col_letter].width
                
                # Copy row heights
                for row_num in source_sheet.row_dimensions:
                    if row_num in source_sheet.row_dimensions:
                        target_sheet.row_dimensions[row_num].height = \
                            source_sheet.row_dimensions[row_num].height
                
                # Copy merged cells
                for merged_cell_range in source_sheet.merged_cells.ranges:
                    target_sheet.merge_cells(str(merged_cell_range))
                
                sheets_copied += 1
            else:
                print(f"  WARNING: No sheet found matching pattern '{pattern}'")
        
        # Save the new workbook
        output_file = os.path.join(output_dir, f'Appendix_{appendix_key}.xlsx')
        wb_new.save(output_file)
        print(f"\n✓ Saved: {output_file}")
        print(f"  Sheets included: {sheets_copied} data sheets + 1 README sheet")
    
    print(f"\n{'='*60}")
    print("PROCESSING COMPLETE")
    print(f"{'='*60}")
    print(f"3 files created in: {output_dir}")
    print("  - Appendix_6-3.xlsx (Rain-fed crops)")
    print("  - Appendix_6-4.xlsx (Gravity irrigation)")
    print("  - Appendix_6-5.xlsx (Drip irrigation)")


if __name__ == "__main__":
    XLSX = "gaez_data/GAEZ4_Appendices.xlsx"
    split_excel_file(XLSX)
    
    # python split_appendices.py "gaez_data/GAEZ4_Appendices.xlsx"
