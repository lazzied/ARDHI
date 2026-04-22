import csv
from dataclasses import dataclass
import json
from pathlib import Path
import re
from typing import Any, List

from engines.report_augmentation.models_and_io import AugmentedLayer


@dataclass
class LogEntry:
    layer:      str   # e.g. "0-20"
    attribute:  str
    value:      Any
    provenance: str   # READ | AUG | CALC | INTERP | HWSD direct
    flag:       str   # free-text detail / warning


def _parse_provenance_and_flag(raw_flag: str) -> tuple[str, str]:
    """
    Extract a short provenance token (READ/AUG/CALC/INTERP/HWSD direct)
    and a cleaned flag string from the raw provenance note.
    """
    tokens = ("READ", "AUG", "CALC", "INTERP", "HWSD direct")
    prov = "UNKNOWN"
    for t in tokens:
        if raw_flag.startswith(t):
            prov = t
            break
    # The flag is everything after the leading "TOKEN | "
    flag = re.sub(r"^(READ|AUG|CALC|INTERP|HWSD direct)\s*\|\s*", "", raw_flag, count=1)
    return prov, flag


class ProvenanceLogger:
    """
    Collects LogEntry objects during pipeline execution and writes them
    to both CSV and JSON when flush() is called.  Fully decoupled from
    the Excel exporter — satisfies SRP.
    """

    def __init__(self) -> None:
        self._entries: List[LogEntry] = []

    def record(self, layer_label: str, augmented: "AugmentedLayer") -> None:
        """Ingest all attributes from one AugmentedLayer."""
        for attr in augmented.values:
            raw_flag = augmented.flags.get(attr, "")
            prov, flag = _parse_provenance_and_flag(raw_flag)
            self._entries.append(LogEntry(
                layer     = layer_label,
                attribute = attr,
                value     = augmented.values[attr],
                provenance= prov,
                flag      = flag,
            ))

    def flush_csv(self, path: str) -> None:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f, fieldnames=["layer", "attribute", "value", "provenance", "flag"]
            )
            writer.writeheader()
            for e in self._entries:
                writer.writerow({
                    "layer":      e.layer,
                    "attribute":  e.attribute,
                    "value":      e.value,
                    "provenance": e.provenance,
                    "flag":       e.flag,
                })
        print(f"✓ Log (CSV)  → {path}")

    def flush_json(self, path: str) -> None:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        records = [
            {
                "layer":      e.layer,
                "attribute":  e.attribute,
                "value":      e.value,
                "provenance": e.provenance,
                "flag":       e.flag,
            }
            for e in self._entries
        ]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        print(f"✓ Log (JSON) → {path}")

    def flush(self, csv_path: str, json_path: str) -> None:
        self.flush_csv(csv_path)
        self.flush_json(json_path)


# ===========================================================================
# VII — EXCEL EXPORTER  (SRP)  — updated: values only, no metadata columns
# ===========================================================================

class ExcelExporter:
    """
    7-sheet workbook.
    Layout (transposed):
      Row 1 : attribute names as column headers
      Row 2 : final processed values
    No provenance/flag columns — those go to the log file.
    """

    def export(self, layers: List[AugmentedLayer], output_path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill    = PatternFill("solid", start_color="2E6B3E")
        value_font     = Font(name="Arial", bold=True, size=10)
        label_font     = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        label_fill     = PatternFill("solid", start_color="5A9E6F")
        center         = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border    = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        for layer in layers:
            sheet_name = f"{layer.layer} ({layer.top_dep}-{layer.bot_dep}cm)"
            ws = wb.create_sheet(title=sheet_name)
            attributes = list(layer.values.keys())

            # Row 1 — label cell + attribute headers
            label_cell = ws.cell(row=1, column=1, value="")
            label_cell.fill   = label_fill
            label_cell.border = thin_border

            for col_idx, attr in enumerate(attributes, start=2):
                c = ws.cell(row=1, column=col_idx, value=attr)
                c.font      = header_font
                c.fill      = header_fill
                c.alignment = center
                c.border    = thin_border

            # Row 2 — "Value" label + data cells
            row_label = ws.cell(row=2, column=1, value="Value")
            row_label.font      = label_font
            row_label.fill      = label_fill
            row_label.alignment = center
            row_label.border    = thin_border

            for col_idx, attr in enumerate(attributes, start=2):
                c = ws.cell(row=2, column=col_idx, value=layer.values[attr])
                c.font      = value_font
                c.alignment = center
                c.border    = thin_border

            # Column widths
            ws.column_dimensions["A"].width = 10
            for col_idx in range(2, len(attributes) + 2):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 22
            ws.freeze_panes = "B2"

        wb.save(output_path)
        print(f"✓ Excel      → {output_path}")
